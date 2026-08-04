"""XLSX (Excel) file adapter.

Uses Pandas with the ``openpyxl`` engine for both reading and writing.
Supports worksheet selection, header row configuration, and sheet
name discovery.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from deskx.processing.interfaces import FileAdapter


class XlsxAdapter(FileAdapter):
    """Read and write XLSX files via openpyxl."""

    @property
    def extensions(self) -> frozenset[str]:
        return frozenset({".xlsx"})

    @property
    def display_name(self) -> str:
        return "Excel (XLSX)"

    # ── Reading ─────────────────────────────────────────────────────

    def read_preview(
        self,
        path: Path,
        max_rows: int = 200,
        **kwargs: Any,
    ) -> pd.DataFrame:
        sheet = kwargs.get("sheet_name", 0)
        header_row = kwargs.get("header_row", 0)
        return pd.read_excel(
            path,
            engine="openpyxl",
            sheet_name=sheet,
            header=header_row,
            nrows=max_rows,
        )

    def read_full(self, path: Path, **kwargs: Any) -> pd.DataFrame:
        sheet = kwargs.get("sheet_name", 0)
        header_row = kwargs.get("header_row", 0)
        return pd.read_excel(
            path,
            engine="openpyxl",
            sheet_name=sheet,
            header=header_row,
        )

    # ── Writing ─────────────────────────────────────────────────────

    def write(
        self,
        df: pd.DataFrame,
        output_path: Path,
        **kwargs: Any,
    ) -> None:
        sheet_name = kwargs.pop("sheet_name", "Sheet1")
        df.to_excel(
            output_path,
            index=False,
            engine="openpyxl",
            sheet_name=sheet_name,
            **kwargs,
        )

    # ── Copy ────────────────────────────────────────────────────────

    def copy_file(self, source: Path, destination: Path) -> None:
        shutil.copy2(source, destination)

    # ── Sheet discovery ─────────────────────────────────────────────

    def get_sheet_names(self, path: Path) -> list[str]:
        """Return the list of worksheet names in the workbook."""
        wb = load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
